from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Font, Alignment

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
driver = webdriver.Chrome(service=service, options=options)

wb = Workbook()
sheet = wb.active

border_style = Side(border_style = 'thin', color = '000000')
item = ['Company name', 'Phone number', 'Website']

for i in range(1, 4):
    sheet.cell(row = 1, column = i).value = item[i-1]
    sheet.cell(row = 1, column = i).border = Border(right = border_style, bottom = border_style)
    sheet.cell(row = 1, column = i).font = Font(name = 'Times New Roman', size = '12')
    sheet.cell(row = 1, column = i).alignment = Alignment(vertical = 'center', horizontal = 'center')

driver.get('https://www.yellowpages.com.pr/en/direcci%C3%B3n/restaurant/Puerto-Rico')
company_names = Find_Elements(driver, By.CLASS_NAME, 'mdc-shape-container')
output = []

start_row = 2
for company_name in company_names:
    try:
        company = company_name.find_element(By.TAG_NAME, 'h2').text
        print(company)
        url = company_name.find_element(By.TAG_NAME, 'a').get_attribute('href')
        output.append({"company_link" : url})
        sheet.cell(row = start_row, column = 1).value = company
        start_row += 1
    except:
        pass

with open('output.json', 'w') as file:
    json.dump(output, file, indent = 4)

start_row = 2
for item_index, item in enumerate(output):
    driver.get(item["company_link"])
    try:
        phone = Find_Element(driver, By.CLASS_NAME, 'phone-header').get_attribute('data-phone-number')
        print(phone)
        sheet.cell(row = start_row, column = 2).value = phone
        website = Find_Element(driver, By.CLASS_NAME, 'company-header-www').find_element(By.TAG_NAME, 'a').get_attribute('href')
        print(website)
        sheet.cell(row = start_row, column = 3).value = website
    except:
        pass
    start_row += 1
    print('done')

for id in range(2, 11):
    driver.get(f'https://www.yellowpages.com.pr/en/direcci%C3%B3n/restaurant/Puerto-Rico/{id}.html')
    company_names = Find_Elements(driver, By.CLASS_NAME, 'mdc-shape-container')
    start_row = 27+25*(id-2)

    with open('output.json', 'r') as file:
        output = json.load(file)
    
    for company_name in company_names:
        try:
            company = company_name.find_element(By.TAG_NAME, 'h2').text
            print(company)
            url = company_name.find_element(By.TAG_NAME, 'a').get_attribute('href')
            output.append({"company_link" : url})
            sheet.cell(row = start_row, column = 1).value = company
            start_row += 1
        except:
            pass
    
    with open('output.json', 'w') as file:
        json.dump(output, file)

    start_row = 27+25*(id-2)
    for item_index, item in enumerate(output):
        driver.get(item["company_link"])
        try:
            phone = Find_Element(driver, By.CLASS_NAME, 'phone-header').get_attribute('data-phone-number')
            print(phone)
            sheet.cell(row = start_row, column = 2).value = phone
            website = Find_Element(driver, By.CLASS_NAME, 'company-header-www').find_element(By.TAG_NAME, 'a').get_attribute('href')
            print(website)
            sheet.cell(row = start_row, column = 3).value = website
        except:
            pass
        start_row += 1
        print('done')
    sleep(0.5)

wb.save('output.xlsx')